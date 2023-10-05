import os
import concurrent.futures
import pandas as pd
from pyrogram.types import Message
from pyrogram import Client
from datetime import datetime
from dotenv import load_dotenv
from openpyxl import load_workbook
from ..service.lancamento_service import lancamento
from ..util.formatador import formatar_valor, formatar_data

load_dotenv()
running = False

def handle_start_lancamento(client: Client, message: Message):
    global running
    if not running:
        # Verifique se a mensagem contém um documento e se o tipo MIME do documento é "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if message.document and (message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
            running = True
            # Quantidade de itens na Pool
            limite_threads = 10

            # Baixe o arquivo XLSX
            file_path = message.download(in_memory=True)
            hora = datetime.now()
            file_name = hora.strftime("%S_%M_%H %Y-%m-%d.log")
            message.reply_text("Preparando arquivo XLSX")
            agente = f"{message.from_user.first_name}.{message.from_user.last_name}"

            # caminho pasta de logs
            diretorio_logs = os.path.join(os.path.dirname(__file__), 'logs')

            # caminho pasta de docs
            diretorio_docs = os.path.join(os.path.dirname(__file__), 'docs')

            # cria pasta de logs em caso de nao existir
            if not os.path.exists(diretorio_logs):
                os.makedirs(diretorio_logs)

            # cria pasta de docs em caso de nao existir
            if not os.path.exists(diretorio_docs):
                os.makedirs(diretorio_docs)
            
            resultados = []
            # Processar o arquivo XLSX conforme necessário
            try:
                try:
                    # Ler o arquivo XLSX usando pandas e especificar a codificação UTF-8
                    # df = pd.read_excel(file_path)
                    workbook = load_workbook(file_path)
                    sheet = workbook['Sheet1']

                    # Converter o dataframe para uma lista de dicionários
                    # lista = df.to_dict(orient='records')
                    lista = {}
                    for row in sheet.iter_rows(values_only=True):
                        key = row[0]  # Suponha que a primeira coluna seja a chave
                        value = row[1]  # Suponha que a segunda coluna seja o valor
                        lista[key] = value


                    # Verificar se a chave 'MK' contém valor NaN
                    lista = [dados for dados in lista if not pd.isna(dados.get('MK'))]

                    # Criar aquivo de log com todos os vencimentos enviados para cancelamento
                    with open(os.path.join(diretorio_docs, file_name), "a") as pedido:
                        for c,arg in enumerate(lista):
                            pedido.write(f"{(c + 1):03};Lançamento;MK:{int(arg.get('MK'))};Negócio:{arg.get('NEGOCIO')};Valor:{arg.get('VALOR')};Descrição:{arg.get('DESCRICAO')};Agente:{agente}\n")
                    
                    # Envia arquivo de docs com todos as solicitações de cancelamento
                    with open(os.path.join(diretorio_docs, file_name), "rb") as enviar_docs:
                        client.send_document(os.getenv("CHAT_ID_ADM"),enviar_docs, caption=f"solicitações {file_name}", file_name=f"solicitações {file_name}")

                    
                    message.reply_text(f"Processando arquivo XLSX de lançamento com {len(lista)} fatura...")

                except pd.errors.ParserError:
                    message.reply_text("O arquivo fornecido não é um arquivo XLSX válido.")
                    running = False
                    return
                
                def executar(arg: dict):
                    if running:
                        try:
                            mk = arg.get("MK")
                            credor = arg.get("CREDOR")
                            vencimento = formatar_data(arg.get("VENCIMENTO"))
                            efetiva = formatar_data(arg.get("EFETIVA"))
                            descricao = arg.get("DESCRICAO")
                            plano_conta: str = arg.get("PLANO_CONTA")
                            combinacao: str = arg.get("COMBINACAO")
                            negocio = arg.get("NEGOCIO")
                            valor = formatar_valor(arg.get("VALOR"))
                            conta: str = arg.get("CONTA")

                            return lancamento(
                                mk = mk,
                                credor = credor,
                                vencimento = vencimento,
                                efetiva = efetiva,
                                descricao = descricao,
                                plano_conta = plano_conta.split(" ")[0],
                                combinacao = combinacao[:9],
                                negocio = negocio,
                                valor = valor,
                                conta = conta,
                                )
                        except Exception as e:
                            print(f'Error executar na função lançamento: MK:{int(arg.get("MK"))} Negócio:{arg.get("NEGOCIO")} Valor:{arg.get("VALOR")} Descrição:{arg.get("DESCRIÇÃO")} {e}')
                    else:
                        message.reply_text(f'Lançamento MK:{int(arg.get("MK"))} Negócio:{arg.get("NEGOCIO")} Valor:{arg.get("VALOR")} Descrição:{arg.get("DESCRIÇÃO")} parado.')
                
                # Criando Pool
                with concurrent.futures.ThreadPoolExecutor(max_workers=limite_threads) as executor:
                    resultados = executor.map(executar, lista)

                # Criar aquivo de log com todos os resultados de cancelamento
                with open(os.path.join(diretorio_logs, file_name), "a") as file:
                    if resultados:
                        for resultado in resultados:
                            file.write(f"{resultado}\n")

                # Envia arquivo de log com todos os resultados de cancelamento
                with open(os.path.join(diretorio_logs, file_name), "rb") as enviar_logs:
                    # message.reply_document(enviar_logs, caption=file_name, file_name=file_name)
                    client.send_document(os.getenv("CHAT_ID_ADM"), enviar_logs, caption=f"resultado {file_name}", file_name=f"resultado {file_name}")

                print("Processo Lançamento concluído.")
                message.reply_text("O arquivo XLSX de lançamento foi processado com sucesso!")
                running = False
                return
            
            except Exception as e:
                print(f"Ocorreu um erro ao processar o arquivo XLSX: {e}")
                message.reply_text("Ocorreu um error ao processar o arquivo XLSX.\nEntre em contato com o Administrador.")
                running = False
                return
        
        else:
            # Responder à mensagem do usuário com uma mensagem de erro
            message.reply_text("Por favor, envie um arquivo XLSX para processar.")
            return
    else:
        message.reply_text("Lançamento em execução.")
        return

def handle_stop_lancamento(client: Client, message: Message):
    global running
    if running:
        running = False
        message.reply_text("Pedido de parada iniciado...")
        return
    else:
        message.reply_text("Lançamento parado")
        return
        
def handle_status_lancamento(client: Client, message: Message):
    global running
    try:
        if running:
            message.reply_text("Lançamento em execução")
            return
        else:
            message.reply_text("Lançamento parado")
            return
    except:
        message.reply_text("Lançamento parado")
        return